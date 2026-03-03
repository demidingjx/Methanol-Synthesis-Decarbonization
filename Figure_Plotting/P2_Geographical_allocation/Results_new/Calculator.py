import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

                                      
base_dir = Path(r"E:\2025_Methanol_synthesis\3_GAMS\P3_Geographical_allocation\Results_new")
out_file = base_dir / "Calculator.xlsx"
sheet_out = "Total_TNS"

                                  
cols = [
    {"year": "2040", "ed_map": {"PEM": "25",  "SOE": "21"}},
    {"year": "2040", "ed_map": {"PEM": "55",  "SOE": "50"}},
    {"year": "2040", "ed_map": {"PEM": "75",  "SOE": "75"}},    
    {"year": "2040", "ed_map": {"PEM": "100", "SOE": "100"}},
    {"year": "2050", "ed_map": {"PEM": "25",  "SOE": "25"}},
    {"year": "2050", "ed_map": {"PEM": "50",  "SOE": "59"}},
    {"year": "2050", "ed_map": {"PEM": "76",  "SOE": "75"}},
    {"year": "2050", "ed_map": {"PEM": "100", "SOE": "100"}},
]

             
row_city_title   = 1
row_city_year    = 2
row_city_ed_pem  = 3
row_city_pem     = 4                               
row_city_ed_soe  = 5
row_city_soe     = 6                               

row_prov_title   = 9
row_prov_year    = 10
row_prov_pem_lab = 11
row_prov_pem     = 12                      
row_prov_soe_ed  = 13
row_prov_soe     = 14                      

row_total_title  = 17
row_total_year   = 18
row_total_pem_ed = 19
row_total_pem    = 20              
row_total_soe_ed = 21
row_total_soe    = 22              

                                        
def read_sums(tech: str, year: str, ed: str):
    f = base_dir / tech / year / f"ED{ed}" / f"TNS_yr_{tech}_{year}_ED{ed}_2s.xlsx"
    df_city = pd.read_excel(f, sheet_name="U_city_Prov_Sum_Total", usecols=["Total_Value"])
    df_prov = pd.read_excel(f, sheet_name="U_prov_region",           usecols=["Value"])
    return df_city["Total_Value"].sum(), df_prov["Value"].sum()

               
city_sum = {"PEM": [None]*8, "SOE": [None]*8}
prov_sum = {"PEM": [None]*8, "SOE": [None]*8}

for j, col in enumerate(cols):
    for tech in ["PEM", "SOE"]:
        year = col["year"]
        ed   = col["ed_map"][tech]
        c_sum, p_sum = read_sums(tech, year, ed)
        city_sum[tech][j] = float(c_sum)
        prov_sum[tech][j] = float(p_sum)

                                                     
if out_file.exists():
    wb = load_workbook(out_file)
    ws = wb[sheet_out] if sheet_out in wb.sheetnames else wb.create_sheet(sheet_out)
else:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_out

       
ws.cell(row=row_city_title, column=1, value="Annual City Transmission")
ws.cell(row=row_city_pem,   column=1, value="PEME")
ws.cell(row=row_city_soe,   column=1, value="SOEC")

ws.cell(row=row_prov_title, column=1, value="Annual Prov Transmission")
ws.cell(row=row_prov_pem_lab, column=1, value="PEME")
ws.cell(row=row_prov_soe,     column=1, value="SOEC")

ws.cell(row=row_total_title, column=1, value="Annual Total Transmission")
ws.cell(row=row_total_pem,   column=1, value="PEME")
ws.cell(row=row_total_soe,   column=1, value="SOEC")

            
years_row = [c["year"] for c in cols]
pem_ed_row = [c["ed_map"]["PEM"] for c in cols]
soe_ed_row = [c["ed_map"]["SOE"] for c in cols]

for j in range(8):
    col_idx = 2 + j         
    ws.cell(row=row_city_year,   column=col_idx, value=years_row[j])
    ws.cell(row=row_city_ed_pem, column=col_idx, value=pem_ed_row[j])
    ws.cell(row=row_city_pem,    column=col_idx, value=city_sum["PEM"][j])
    ws.cell(row=row_city_ed_soe, column=col_idx, value=soe_ed_row[j])
    ws.cell(row=row_city_soe,    column=col_idx, value=city_sum["SOE"][j])

              
for j in range(8):
    col_idx = 2 + j
    ws.cell(row=row_prov_year,   column=col_idx, value=years_row[j])
    ws.cell(row=row_prov_pem_lab, column=col_idx, value=pem_ed_row[j])                        
    ws.cell(row=row_prov_pem,    column=col_idx, value=prov_sum["PEM"][j])         
    ws.cell(row=row_prov_soe_ed, column=col_idx, value=soe_ed_row[j])
    ws.cell(row=row_prov_soe,    column=col_idx, value=prov_sum["SOE"][j])         

               
for j in range(8):
    col_idx = 2 + j
    ws.cell(row=row_total_year,   column=col_idx, value=years_row[j])
    ws.cell(row=row_total_pem_ed, column=col_idx, value=pem_ed_row[j])
    ws.cell(row=row_total_soe_ed, column=col_idx, value=soe_ed_row[j])

                                
    pem_total = (city_sum["PEM"][j] or 0) + (prov_sum["PEM"][j] or 0)
    soe_total = (city_sum["SOE"][j] or 0) + (prov_sum["SOE"][j] or 0)
    ws.cell(row=row_total_pem, column=col_idx, value=pem_total)
    ws.cell(row=row_total_soe, column=col_idx, value=soe_total)

          
for j in range(1, 6):
    ws.column_dimensions[get_column_letter(j)].width = 16

wb.save(out_file)
print(f"已写入：{out_file} -> {sheet_out}")
